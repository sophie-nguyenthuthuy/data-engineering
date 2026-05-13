"""Excel (.xlsx) source adapter.

Optional dependency on :mod:`openpyxl`. We import it lazily inside
:meth:`ExcelSource.fetch` so that callers who never touch an Excel
source pay no import cost — and CI without the ``[excel]`` extra still
loads the package's ``__init__`` cleanly.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from collections.abc import Iterator

from msc.sources.base import Record, Source, SourceError


@dataclass
class ExcelSource(Source):
    """Local-file ``.xlsx`` source backed by openpyxl."""

    path: Path
    sheet: str | None = None
    dataset: str = ""
    id_column: str | None = None
    kind: str = "excel"

    def __post_init__(self) -> None:
        if not isinstance(self.path, Path):
            self.path = Path(self.path)
        if not self.dataset:
            self.dataset = self.path.stem
        super().__post_init__()

    def fetch(self) -> Iterator[Record]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - exercised via missing extra
            raise SourceError(
                "openpyxl is required for ExcelSource; "
                "install with `pip install multi-source-collector[excel]`"
            ) from exc

        if not self.path.exists():
            raise SourceError(f"Excel file not found: {self.path}")

        wb = load_workbook(filename=str(self.path), read_only=True, data_only=True)
        try:
            ws = wb[self.sheet] if self.sheet else wb.active
            if ws is None:
                raise SourceError(f"Excel workbook has no active sheet: {self.path}")
            rows = ws.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                return
            if not header:
                return
            headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(header)]
            if self.id_column and self.id_column not in headers:
                raise SourceError(f"id_column {self.id_column!r} not in sheet header {headers}")
            for i, row in enumerate(rows, start=1):
                fields = {headers[j]: row[j] for j in range(min(len(headers), len(row)))}
                source_id = (
                    str(fields[self.id_column]) if self.id_column is not None else f"row-{i}"
                )
                yield Record(source_id=source_id, fields=fields)
        finally:
            wb.close()


__all__ = ["ExcelSource"]

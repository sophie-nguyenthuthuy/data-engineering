"""Write SBV report DataFrames to Excel (xlsx) and CSV with proper formatting."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd

from sbv_reporting.utils.config import get_config


REPORT_META = {
    "BCGD":     {"title": "BÁO CÁO GIAO DỊCH NGÀY",              "sheet": "BCGD"},
    "B01-TCTD": {"title": "BÁO CÁO SỐ DƯ VÀ KHỐI LƯỢNG GIAO DỊCH", "sheet": "B01-TCTD"},
    "BCGDLN":   {"title": "BÁO CÁO GIAO DỊCH GIÁ TRỊ LỚN",       "sheet": "BCGDLN"},
    "BCGDNS":   {"title": "BÁO CÁO GIAO DỊCH ĐÁNG NGỜ",           "sheet": "BCGDNS"},
}


class ReportWriter:
    def __init__(self, output_dir: str | Path | None = None):
        self.cfg = get_config()
        base = Path(output_dir or "data/output/reports")
        base.mkdir(parents=True, exist_ok=True)
        self.output_dir = base

    # ------------------------------------------------------------------
    def write_excel(
        self,
        reports: dict[str, pd.DataFrame],
        run_id: str,
        report_date: str | None = None,
    ) -> Path:
        """Write all reports as sheets in a single xlsx file."""
        date_str = (report_date or datetime.today().strftime("%Y%m%d")).replace("/", "")
        filename = self.output_dir / f"SBV_REPORT_{run_id}_{date_str}.xlsx"

        inst = self.cfg["reporting"]

        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            for code, df in reports.items():
                meta = REPORT_META.get(code, {"title": code, "sheet": code[:31]})
                sheet = meta["sheet"]

                # Header block (rows 1-5)
                header_df = pd.DataFrame({
                    "": [
                        f"TỔ CHỨC TÍN DỤNG: {inst['institution_name']}",
                        f"MÃ ĐỊNH CHẾ: {inst['institution_code']}  |  SWIFT: {inst.get('swift_code', '')}",
                        meta["title"],
                        f"Ngày báo cáo: {date_str}  |  Run ID: {run_id}",
                        "",
                    ]
                })
                header_df.to_excel(writer, sheet_name=sheet, index=False, header=False, startrow=0)

                if df.empty:
                    empty_note = pd.DataFrame({"GHI CHÚ": ["Không có dữ liệu trong kỳ báo cáo"]})
                    empty_note.to_excel(writer, sheet_name=sheet, index=False, startrow=6)
                else:
                    df.to_excel(writer, sheet_name=sheet, index=False, startrow=6)
                    self._style_sheet(writer.sheets[sheet], df)

        return filename

    def write_csv(
        self,
        reports: dict[str, pd.DataFrame],
        run_id: str,
        report_date: str | None = None,
    ) -> dict[str, Path]:
        """Write each report as a separate CSV file."""
        date_str = (report_date or datetime.today().strftime("%Y%m%d")).replace("/", "")
        paths: dict[str, Path] = {}
        encoding = self.cfg["sbv_formats"]["encoding"]

        for code, df in reports.items():
            fname = self.output_dir / f"{code}_{run_id}_{date_str}.csv"
            df.to_csv(fname, index=False, encoding=encoding)
            paths[code] = fname

        return paths

    # ------------------------------------------------------------------
    @staticmethod
    def _style_sheet(ws, df: pd.DataFrame) -> None:
        """Apply basic column widths to the worksheet."""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            header_row = 7  # 1-indexed; data starts row 7
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=10)

            for col_idx, col_name in enumerate(df.columns, start=1):
                cell = ws.cell(row=header_row, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

                col_letter = get_column_letter(col_idx)
                max_width = max(len(str(col_name)), 12)
                ws.column_dimensions[col_letter].width = min(max_width + 4, 40)
        except ImportError:
            pass  # openpyxl styles are optional
